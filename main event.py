from selenium import webdriver
from selenium.webdriver.chrome.service import Service 
from selenium.webdriver.common.by import By
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import keyboard
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#hier zorgen we er voor dat we in het excel bestand kunnen kleuren. 
#------------------ Kleur
# Laad het bestaande Excel-bestand
wb = load_workbook("Actieve_Dimona.xlsx")
ws = wb.active  # Werk met het actieve werkblad

# Definieer kleuren
groen = PatternFill(start_color="00B050",  # Groen
                         end_color="00B050",
                         fill_type="solid")

rood = PatternFill(start_color="FF0000",  # Rood
                       end_color="FF0000",
                       fill_type="solid")

geel = PatternFill(start_color="FFFF00",  # Geel
                          end_color="FFFF00",
                          fill_type="solid")

#------------------

#maakt een file waar de namen inkomen
f = open("Namen_Medewerkers.txt", "w")

# Variabelen
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)
DimonaBestand = "Actieve_Dimona.xlsx"
loon = ""
inlog = ""
ww = ""
nummer_medewerker = 1

driver.get("https://my.tentoo.be/login")

# Variabelen voor html naam van invulveld
T_Medewerker = "freelancer-autocomplete"
T_Omschrijving_Opdracht = "Productionname"
T_Plaats = "Location_0"
T_Functie = "Function"
T_Loonbedarg = "Feeamount"
T_Begin_Datum = "moetnog"
T_Eind_Datum = "moetnog2"
Type_gebruikersnaam = "text"
Type_wachtwoord = "password"

# Lijsten maken voor de data
Lijst_Namen = []
Lijst_Plaats = []
Lijst_Uren = []

# Hier lezen we het Excel bestand
D_Naam = pd.read_excel(DimonaBestand, sheet_name=0, usecols=["Medewerker"]).squeeze()
D_Plaats = pd.read_excel(DimonaBestand, sheet_name=0, usecols=["Plaats"]).squeeze()
D_Uren = pd.read_excel(DimonaBestand, sheet_name=0, usecols=["Uren"]).squeeze()

# Data aan lijst toevoegen.
Lijst_Namen.extend(D_Naam.tolist())
Lijst_Plaats.extend(D_Plaats.tolist())
Lijst_Uren.extend(D_Uren.tolist())

time.sleep(3)
# Inloggen
WebDriverWait(driver, 10).until(
    EC.url_contains("https://my.tentoo.be/login")
)
gebruikersnaam_veld = driver.find_element(By.CLASS_NAME, "form-control")
wachtwoord_veld = driver.find_element(By.XPATH, "//input[@type='password' and @placeholder='Wachtwoord']")    
gebruikersnaam_veld.send_keys(inlog)
wachtwoord_veld.send_keys(ww)

# Wacht tot het aanmeldknop-element zichtbaar is
aanmelden_knop = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, "btn-primary"))
)

# Wacht tot het preloader-element verdwijnt
WebDriverWait(driver, 10).until(
    EC.invisibility_of_element_located((By.CLASS_NAME, "preloader"))
)

# Klik op de knop
aanmelden_knop.click()

# Wacht tot de inlogactie is voltooid
WebDriverWait(driver, 10).until(
    EC.url_contains("https://my.tentoo.be/")
)
        


        
def run_loop():

        
        
    for naam, plaats, uur in zip(Lijst_Namen, Lijst_Plaats, Lijst_Uren):
        
        global nummer_medewerker
        driver.get("https://my.tentoo.be/jobsheets/new")
        nummer_medewerker += 1
        try:
            
            inputveld_freelancer = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.ID, "freelancer-autocomplete"))
            )
            inputveld_freelancer.send_keys(naam)

            knop_werknemer = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//ul[@class='dropdown-menu ng-isolate-scope']/li/a"))
            )
            knop_werknemer.click()
        except:
            
            print( naam + "--------- naam niet gevonden/ medewerker geblokkeerd" + str(nummer_medewerker))
            f.write(str(nummer_medewerker)+ naam + "--------- naam niet gevonden/ medewerker geblokkeerd \n" )
            for col in range(1, 14):     
                ws.cell(row= nummer_medewerker, column=col).fill = rood
            wb.save("Actieve_Dimona.xlsx")
            f.flush()
            continue
        try:
            
            inputveld_omschrijving_opdracht = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.NAME, "Productionname"))
            )
            inputveld_omschrijving_opdracht.click()
            inputveld_omschrijving_opdracht.send_keys("Deur aan deur werving")

            inputveld_plaats = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.NAME, "Location_0"))
            )
            inputveld_plaats.send_keys(plaats)

            inputveld_bedrag = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.NAME, "Feeamount"))
            )
            inputveld_bedrag.send_keys(loon)

            time.sleep(1)
            agendadruk = WebDriverWait(driver,5).until(
                EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/section/div/ui-view/div/form/div/div/div[1]/job-sheet-editor-details/div[1]/div[5]/div[2]/div[1]/div[1]/span" ))
                )
            agendadruk.click()
            
        except:
            print( naam + "---------- error bij vinden van agenda of daarvoor bij invullen van informatie" + str(nummer_medewerker))
            f.write(   str(nummer_medewerker)+  naam +"---------- error bij vinden van agenda of daarvoor bij invullen van informatie\n" )
            for col in range(1, 14):     
                ws.cell(row= nummer_medewerker, column=col).fill = rood
            wb.save("Actieve_Dimona.xlsx")
            
            f.flush()
        


        try:
            
            
            # Zoek alle cellen met een 'lipje' en of het vandaag is
            elementen_met_lipje_en_vandaag = driver.find_elements(
                By.XPATH,
                "//td[contains(@class, 'datepicker-marked')]//button[contains(@class, 'active')]//span[contains(@class, 'text-info')]"
            )
            
            # Als er elementen zijn gevonden, betekent dit dat de medewerker al is ingeplant
            if elementen_met_lipje_en_vandaag:
                print(naam + "-------------- Medewerker al ingeplant (geel) " + str(nummer_medewerker))
                f.write(str(nummer_medewerker) + naam + "---------- Medewerker al ingeplant (geel)\n")
                for col in range(1, 14):
                    ws.cell(row=nummer_medewerker, column=col).fill = geel
                wb.save("Actieve_Dimona.xlsx")
                f.flush()
                continue
                
            
                
            dagdruk = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//button[contains(@class, 'active')]//span[contains(@class, 'text-info')]"))
            )
                
            dagdruk.click()

            
            
            confirm = WebDriverWait(driver,5).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='ngdialog1']//div[contains(@class, 'text-right')]//button[normalize-space()='Accepteren']"))
                )
            confirm.click()
                
            time.sleep(1)
            dagdruk2 = WebDriverWait(driver,10).until(
            EC.visibility_of_element_located((By.XPATH, "//button[contains(@class, 'active')]//span[contains(@class, 'text-info')]")) 
            )
            dagdruk2.click()
            

        except :
            print( naam + "-------------- ROOD kan niet vinden in agenda)" + str(nummer_medewerker))
            f.write(   str(nummer_medewerker)+  naam +"---------- error bij vinden van agenda of daarvoor bij invullen van informatie\n" )
            for col in range(1, 14):     
                ws.cell(row= nummer_medewerker, column=col).fill = rood
            wb.save("Actieve_Dimona.xlsx")
            f.flush()
            continue
                   
        try:
            inputveld_uren = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@ng-model='item.Hours']"))
            )
            inputveld_uren.send_keys(uur)
        except:
            continue

        time.sleep(1)
            
        Bevestigingknop = WebDriverWait(driver,5).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id=\"jobsheetForm\"]/div/div/div[1]/job-sheet-editor-footer/div[3]/div[2]/div[1]/span[1]/button" ))
            )
        Bevestigingknop.click()


        try:
            bevestigingknop2 = WebDriverWait(driver,12).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/section/div/ui-view/div/form/div[2]/div[2]/div/div/button[1]"))
            )
            bevestigingknop2.click()

        except:
            print( naam + "----------- error kan niet doordrukken laatste acceptatie" + str(nummer_medewerker))
            f.write(  str(nummer_medewerker)+ naam + "----------- error kan niet doordrukken laatste acceptatie\n" )
            f.flush()
            continue
            
             
        WebDriverWait(driver, 60).until(
        EC.url_contains("https://my.tentoo.be/jobsheets/list/2"))

        #Dubbel check    
        naam_op_pagina = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/section/div/ui-view/div/div[4]/div[1]/div[3]/div[1]/table/tbody/tr[1]/td[3]"))
        )
        #kleine sleep zodat het meer tijd heeft om te laden
        time.sleep(4)
                
        naam_gevonden = naam_op_pagina.text.strip()
        if naam_gevonden == naam:           
            print( naam + " contract is aangemaakt (groen)" + str(nummer_medewerker))
            f.write( str(nummer_medewerker) +naam + " contract is aangemaakt (groen) \n" )
            f.write(   str(nummer_medewerker)+  naam +"---------- error bij vinden van agenda of daarvoor bij invullen van informatie\n" )
            for col in range(1, 14):     
                ws.cell(row= nummer_medewerker, column=col).fill = groen
            wb.save("Actieve_Dimona.xlsx")
            f.flush()
        else:
            print("ERROR! MEDEDEWERKER NIET GEVONDEN IN DUBBEL CHECK! gaat om:" + str(naam));
            f.write(   str(nummer_medewerker)+  naam +"---------- error bij vinden van agenda of daarvoor bij invullen van informatie\n" )
            for col in range(1, 14):     
                ws.cell(row= nummer_medewerker, column=col).fill = rood
            wb.save("Actieve_Dimona.xlsx")
        
            f.write(  str(nummer_medewerker)+ naam + "!DUBBEl CHECK ERROR! \n" )
            f.flush()
            

while True:
     if keyboard.is_pressed('space'):
         run_loop()
    

